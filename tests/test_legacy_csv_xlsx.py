from pathlib import Path
from legacy.csv_gen import generate_csv, HEADERS
from legacy.xlsx_export import csv_to_xlsx
import openpyxl

def test_generate_csv(tmp_path: Path):
    p = tmp_path / "t.csv"
    generate_csv(p, rows=5)
    text = p.read_text(encoding="utf-8")
    assert ",".join(HEADERS) in text
    assert len(text.strip().splitlines()) == 6  # header + 5 rows

def test_csv_to_xlsx(tmp_path: Path):
    csvp = tmp_path / "t.csv"
    xlsxp = tmp_path / "t.xlsx"
    generate_csv(csvp, rows=3)
    csv_to_xlsx(csvp, xlsxp)

    wb = openpyxl.load_workbook(xlsxp)
    ws = wb.active
    assert ws.max_row == 4  # header + 3 rows
    assert ws.max_column == len(HEADERS) + 1  # + exported_at
