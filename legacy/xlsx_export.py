from pathlib import Path
import csv
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def csv_to_xlsx(csv_path: Path, xlsx_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "data"

    with csv_path.open("r", encoding="utf-8") as f:
        r = csv.reader(f)
        headers = next(r)
        ws.append(headers + ["exported_at"])  # подстановка значений (дата/время)
        exported_at = datetime.now(timezone.utc).isoformat()

        for row in r:
            ws.append(row + [exported_at])

    # autosize
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(40, max(10, max_len + 2))

    wb.save(xlsx_path)
